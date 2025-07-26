"""
表格导出器 - 支持多种格式的表格输出
"""
import csv
import json
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime

# 可选依赖
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from tabulate import tabulate
    HAS_TABULATE = True
except ImportError:
    HAS_TABULATE = False

logger = logging.getLogger(__name__)


class TableExporter:
    """表格导出器"""
    
    def __init__(self):
        """初始化导出器"""
        self.supported_formats = ["console", "csv", "json", "html"]
        
        if HAS_PANDAS:
            self.supported_formats.extend(["excel", "parquet"])
        if HAS_OPENPYXL:
            self.supported_formats.append("xlsx")
        if HAS_TABULATE:
            self.supported_formats.append("markdown")
    
    def export(self, 
               data: List[Dict[str, Any]], 
               format_type: str = "console",
               output_path: Optional[str] = None,
               **kwargs) -> str:
        """
        导出表格数据
        
        Args:
            data: 表格数据列表
            format_type: 输出格式
            output_path: 输出文件路径
            **kwargs: 额外参数
            
        Returns:
            导出结果信息
        """
        if not data:
            return "没有数据可导出"
        
        if format_type not in self.supported_formats:
            raise ValueError(f"不支持的格式: {format_type}. 支持的格式: {self.supported_formats}")
        
        try:
            if format_type == "console":
                return self._export_console(data, **kwargs)
            elif format_type == "csv":
                return self._export_csv(data, output_path, **kwargs)
            elif format_type == "json":
                return self._export_json(data, output_path, **kwargs)
            elif format_type == "html":
                return self._export_html(data, output_path, **kwargs)
            elif format_type == "excel" and HAS_PANDAS:
                return self._export_excel_pandas(data, output_path, **kwargs)
            elif format_type == "xlsx" and HAS_OPENPYXL:
                return self._export_xlsx(data, output_path, **kwargs)
            elif format_type == "markdown" and HAS_TABULATE:
                return self._export_markdown(data, output_path, **kwargs)
            elif format_type == "parquet" and HAS_PANDAS:
                return self._export_parquet(data, output_path, **kwargs)
            else:
                raise ValueError(f"格式 {format_type} 需要额外的依赖库")
                
        except Exception as e:
            logger.error(f"导出失败: {e}")
            raise
    
    def _export_console(self, data: List[Dict[str, Any]], **kwargs) -> str:
        """导出到控制台"""
        if not HAS_TABULATE:
            # 简单的文本表格
            return self._export_simple_table(data)
        
        # 使用tabulate美化输出
        headers = list(data[0].keys()) if data else []
        rows = [[row.get(header, "") for header in headers] for row in data]
        
        table_format = kwargs.get("table_format", "grid")
        max_col_width = kwargs.get("max_col_width", 50)
        
        # 截断过长的内容
        for row in rows:
            for i, cell in enumerate(row):
                if isinstance(cell, str) and len(cell) > max_col_width:
                    row[i] = cell[:max_col_width-3] + "..."
        
        table_str = tabulate(rows, headers=headers, tablefmt=table_format)
        print(table_str)
        
        return f"已在控制台显示 {len(data)} 行数据"
    
    def _export_simple_table(self, data: List[Dict[str, Any]]) -> str:
        """简单的文本表格输出"""
        if not data:
            return "无数据"
        
        headers = list(data[0].keys())
        
        # 计算列宽
        col_widths = {}
        for header in headers:
            col_widths[header] = max(
                len(header),
                max(len(str(row.get(header, ""))) for row in data)
            )
            # 限制最大宽度
            col_widths[header] = min(col_widths[header], 50)
        
        # 构建表格
        lines = []
        
        # 表头
        header_line = " | ".join(header.ljust(col_widths[header]) for header in headers)
        lines.append(header_line)
        lines.append("-" * len(header_line))
        
        # 数据行
        for row in data:
            row_line = " | ".join(
                str(row.get(header, "")).ljust(col_widths[header])[:col_widths[header]]
                for header in headers
            )
            lines.append(row_line)
        
        table_str = "\n".join(lines)
        print(table_str)
        
        return f"已在控制台显示 {len(data)} 行数据"
    
    def _export_csv(self, data: List[Dict[str, Any]], output_path: Optional[str], **kwargs) -> str:
        """导出为CSV格式"""
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"news_export_{timestamp}.csv"
        
        encoding = kwargs.get("encoding", "utf-8-sig")  # 支持Excel打开中文
        
        with open(output_path, 'w', newline='', encoding=encoding) as csvfile:
            if data:
                fieldnames = list(data[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)
        
        return f"已导出 {len(data)} 行数据到 {output_path}"
    
    def _export_json(self, data: List[Dict[str, Any]], output_path: Optional[str], **kwargs) -> str:
        """导出为JSON格式"""
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"news_export_{timestamp}.json"
        
        export_data = {
            "export_time": datetime.now().isoformat(),
            "total_count": len(data),
            "data": data
        }
        
        with open(output_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(export_data, jsonfile, ensure_ascii=False, indent=2)
        
        return f"已导出 {len(data)} 行数据到 {output_path}"
    
    def _export_html(self, data: List[Dict[str, Any]], output_path: Optional[str], **kwargs) -> str:
        """导出为HTML格式"""
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"news_export_{timestamp}.html"
        
        # 构建HTML表格
        html_content = self._build_html_table(data, **kwargs)
        
        with open(output_path, 'w', encoding='utf-8') as htmlfile:
            htmlfile.write(html_content)
        
        return f"已导出 {len(data)} 行数据到 {output_path}"
    
    def _build_html_table(self, data: List[Dict[str, Any]], **kwargs) -> str:
        """构建HTML表格"""
        if not data:
            return "<p>无数据</p>"
        
        title = kwargs.get("title", "新闻筛选结果")
        
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        .url {{ max-width: 200px; word-break: break-all; }}
        .content {{ max-width: 300px; }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    <p>导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
    <p>总计: {len(data)} 条记录</p>
    
    <table>
        <thead>
            <tr>
"""
        
        # 表头
        headers = list(data[0].keys())
        for header in headers:
            html += f"                <th>{header}</th>\n"
        
        html += """            </tr>
        </thead>
        <tbody>
"""
        
        # 数据行
        for row in data:
            html += "            <tr>\n"
            for header in headers:
                value = str(row.get(header, ""))
                css_class = ""
                
                if header == "链接":
                    css_class = "url"
                    if value and value.startswith("http"):
                        value = f'<a href="{value}" target="_blank">{value}</a>'
                elif header in ["原文全文", "中文摘要"]:
                    css_class = "content"
                    if len(value) > 200:
                        value = value[:200] + "..."
                
                html += f'                <td class="{css_class}">{value}</td>\n'
            html += "            </tr>\n"
        
        html += """        </tbody>
    </table>
</body>
</html>"""
        
        return html
    
    def _export_xlsx(self, data: List[Dict[str, Any]], output_path: Optional[str], **kwargs) -> str:
        """使用openpyxl导出Excel格式"""
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"news_export_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "新闻筛选结果"
        
        if not data:
            wb.save(output_path)
            return f"已创建空文件 {output_path}"
        
        headers = list(data[0].keys())
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=str(value))
        
        # 调整列宽
        try:
            from openpyxl.utils import get_column_letter
            for col_num, col in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            print(f"⚠️ 调整列宽失败: {e}")
        
        wb.save(output_path)
        return f"已导出 {len(data)} 行数据到 {output_path}"
    
    def _export_excel_pandas(self, data: List[Dict[str, Any]], output_path: Optional[str], **kwargs) -> str:
        """使用pandas导出Excel格式"""
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"news_export_{timestamp}.xlsx"
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='新闻筛选结果', index=False)
            
            # 格式化工作表
            worksheet = writer.sheets['新闻筛选结果']
            try:
                from openpyxl.utils import get_column_letter
                for col_num, column in enumerate(worksheet.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(col_num)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            except Exception as e:
                print(f"⚠️ 调整列宽失败: {e}")
        
        return f"已导出 {len(data)} 行数据到 {output_path}"
    
    def _export_markdown(self, data: List[Dict[str, Any]], output_path: Optional[str], **kwargs) -> str:
        """导出为Markdown格式"""
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"news_export_{timestamp}.md"
        
        if not data:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("# 新闻筛选结果\n\n无数据\n")
            return f"已创建空文件 {output_path}"
        
        headers = list(data[0].keys())
        rows = [[str(row.get(header, "")) for header in headers] for row in data]
        
        markdown_table = tabulate(rows, headers=headers, tablefmt="pipe")
        
        content = f"""# 新闻筛选结果

导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
总计: {len(data)} 条记录

{markdown_table}
"""
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        return f"已导出 {len(data)} 行数据到 {output_path}"
    
    def _export_parquet(self, data: List[Dict[str, Any]], output_path: Optional[str], **kwargs) -> str:
        """导出为Parquet格式"""
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"news_export_{timestamp}.parquet"
        
        df = pd.DataFrame(data)
        df.to_parquet(output_path, index=False)
        
        return f"已导出 {len(data)} 行数据到 {output_path}"
    
    def get_supported_formats(self) -> List[str]:
        """获取支持的导出格式"""
        return self.supported_formats.copy()


# 便捷函数
def export_table(data: List[Dict[str, Any]], 
                format_type: str = "console",
                output_path: Optional[str] = None,
                **kwargs) -> str:
    """
    导出表格数据的便捷函数
    
    Args:
        data: 表格数据
        format_type: 导出格式
        output_path: 输出路径
        **kwargs: 额外参数
        
    Returns:
        导出结果信息
    """
    exporter = TableExporter()
    return exporter.export(data, format_type, output_path, **kwargs)
